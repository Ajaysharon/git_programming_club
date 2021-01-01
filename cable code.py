from openpyxl import load_workbook
name='MonthlyBill.xlsx'
wb=load_workbook(filename=""+name+"")
sheet=wb.active
l=[]
max_col=sheet.max_row
sheet.insert_cols(11)
for i in range(1,max_col+1):
    cells=sheet.cell(row=i,column=10)
    l.append(cells.value)
    
L=[]

for i in range(len(l)):
    sp=l[i].split(",")
    t=0
   
    for j in range(len(sp)):
        T=[]
        for i in range(len(sp[j])):
            T.append(sp[j][i])
   
        if T[0].isspace()==True:
            T.pop(0)
    
        if T[-1].isspace()==True:
            T.pop(-1) 
        c=''
        for x in T:
            c+=x
        
        if c=='classic':
            t=t+(218.20)
        elif c=='TCCL BASIC FTA 170 Channel':
            t=t+(170)
        elif c=='TCCL CLASSIC SD':
            t=t+(218.20)
        elif c=='TCCL CL':
            t=t+(218.20)
        elif c=='TCCL CLASSIC':
            t=t+(218.20)
        elif c=='TCCL BASIC TAMIL SD PACK':
            t=t+(195)
        elif c=='suntv':
            t=t+(19+(19*0.18))
        elif c=='SUN TV':
            t=t+(19+(19*0.18))
        elif c=='ASIANET':
            t=t+(19+(19*0.18))
        elif c=='DISCOVERY TAMIL FAMILY PACK':
            t=t+(8+(8*0.18))
        elif c=='SURYA TV':
            t=t+(12+(12*0.18))
        elif c=='ZEE PRIME MOVIE PACK TAMIL SD':
            t=t+(10+(10*0.18))
        elif c=='ASIANET PLUS':
            t=t+(5+(5*0.18))
        elif c=='ASIANET MOVIES':
            t=t+(15+(15*0.18))
        elif c=='CHUTTI TV':
            t=t+(6+(6*0.18))
        elif c=='JAYA PLUS':
            t=t+(0.5+(0.5*0.18))
        elif c=='TURNER KIDS PACK':
            t=t+(4.25+(4.25*0.18))
        elif c=='DISNEY KIDS PACK':
            t=t+(10+(10*0.18))
        elif c=='FLOWERS TV':
            t=t+(10+(10*0.18))
        elif c=='ASIANET HD':
            t=t+(19+(19*0.18))
        elif c=='FOX LIFE HD':
            t=t+(1+(1*0.18))
        elif c=='DISCOVERY HD WORLD':
            t=t+(6+(6*0.18))
        elif c=='HUNGAMA TV':
            t=t+(6+(6*0.18))
        elif c=='VIJAY TV':
            t=t+(17+(17*0.18))
        elif c=='ICAST TAMIL BUDGET':
            t=t+(8+(8*0.18))
        elif c=='JAYA TAMIL PACK 1':
            t=t+(6+(6*0.18))
        elif c=='NEWS18 TAMIL NADU':
            t=t+(0.10+(0.10*0.18))
        elif c=='NICK':
            t=t+(6+(6*0.18))
        elif c=='STAR SPORTS 1 TAMIL':
            t=t+(17+(17*0.18))
        elif c=='STAR TAMIL MALAYALAM VALUE':
            t=t+(49+(49*0.18))
        elif c=='SUN NEWS':
            t=t+(1+(1*0.18))
        elif c=='TCCL TAMIL MALAYALAM SD PACK':
            t=t+(260)
        elif c=='COLORS TAMIL':
            t=t+(3+(3*0.18))
        elif c=='SUN LIFE':
            t=t+(9+(9*0.18))
        elif c=='ADITHYA TV':
            t=t+(9+(9*0.18))
        elif c=='OM TV':
            t=t+(1+(1*0.18))
        elif c=='STAR TAMIL VALUE':
            t=t+(25+(25*0.18))
        elif c=='TRAVEL XP TAMIL':
            t=t+(1.50+(1.50*0.18))
        elif c=='BINDASS TV':
            t=t+(0.10+(0.10*0.18))
        elif c=='MEGA 24':
            t=t+(1+(1*0.18))
        elif c=='NDTV GOOD TIMES':
            t=t+(1.50+(1.50*0.18))
        elif c=='ZING':
            t=t+(0.10+(0.10*0.18))
        elif c=='MTV BEATS HD':
            t=t+(1+(1*0.18))
        elif c=='VH1':
            t=t+(1+(1*0.18))
        elif c=='MTV BEATS':
            t=t+(0.10+(0.10*0.18))
        elif c=='M TV':
            t=t+(3+(3*0.18))
        elif c=='DISCOVERY TAMIL FAMILY PACK':
            t=t+(8+(8*0.18))
        elif c=='SUN MUSIC':
            t=t+(6+(6*0.18))
        elif c=='SURYA MOVIES':
            t=t+(11+(11*0.18))
        elif c=='FOX LIFE':
            t=t+(1+(1*0.18))
        elif c=='ZEE BUSINESS':
            t=t+(0.10+(0.10*0.18))
        elif c=='CARTOON NETWORK':
            t=t+(4.25+(4.25*0.18))
        elif c=='DISNEY CHANNEL':
            t=t+(8+(8*0.18))
        elif c=='VIJAY SUPER':
            t=t+(2+(2*0.18))
        elif c=='SONY HAPPY INDIA ENGLISH SD':
            t=t+(12+(12*0.18))
        elif c=='TVTN NEWS BOUQUET':
            t=t+(1+(1*0.18))
        elif c=='TCCL CLASSIC PACK 1 YEAR':
            t=t+230
        elif c=='RAJ TAMIL PACK':
            t=t+(4.89+(4.89*0.18))
        elif c=='TURNER FAMILY PACK':
            t=t+(10+(10*0.18))
        elif c=='MEGA TAMIL PACK 1':
            t=t+(3.60+(3.60*0.18))
        elif c=='TIMES VALUE PACK':
            t=t+(13+(13*0.18))
        elif c=='JAYA TAMIL PACK 1':
            t=t+(6+(6*0.18))
        elif c=='SUN TAMIL BASIC':
            t=t+(40+(40*0.18))   
            
            
            
        elif c=='UTV BINDASS':
            t=t+(1+(1*0.18)) 
        elif c=='NDTV GOODTIMES':
            t=t+(1.50+(1.50*0.18))
        elif c=='JAYA TAMIL PACK':
            t=t+(6+(6*0.18))
            
        elif c=='ZEE TAMIL':
            t=t+(10+(10*0.18))        
            
        elif c=='DISCOVERY TAMIL':
            t=t+(4+(4*0.18))    
        elif c=='ANIMAL PLANET':
            t=t+(2+(2*0.18))    
        
            
        elif c=='SONY YAY':
            t=t+(2+(2*0.18))
        elif c=='SAI TV':
            t=t+0
        elif c=='SONY TEN 1':
            t=t+(19+(19*0.18))
        elif c=='SONY YAY':
            t=t+(2+(2*0.18))    
        elif c=='BBC WORLD NEWS':
            t=t+(1+(1*0.18))
        elif c=='POGO':
            t=t+(4.25+(4.25*0.18))
        elif c=='SUN TAMIL BASIC':
            t=t+(40+(40*0.18))
        else:
            k=float(input("enter the value of :"+c+": without tax: "))
            t=t+(k+(k*0.18))
    L.append(round(t,2))
for k in range(len(L)):
    if L[k]==218.2:
        L[k]=230
    elif L[k]<=169:
        L[k]='DA'
    elif L[k]==604.72:
        L[k]=230
    
    
    
for x in range(len(L)):
    B=sheet.cell(row=x+1,column=11)
    B.value=L[x]
wb.save(filename=""+name+"")

from openpyxl import load_workbook
wb=load_workbook(filename=""+name+"")
sheet=wb.active
l1=[]
max_col=sheet.max_row
for x in range(1,max_col+1):
    cells=sheet.cell(row=x,column=2)
    l1.append(cells.value)
L1=[]
for y in range(len(l1)):
    if l1[y]=='C1424009':
       L1.append(1)
    elif l1[y]=='C1432790':
       L1.append(2)
    elif l1[y]=='C1432919':
       L1.append(3)
    elif l1[y]=='C1432920':
       L1.append(4)
    elif l1[y]=='C1432921':
       L1.append(5)      
    elif l1[y]=='C1432922':
       L1.append(6)
    elif l1[y]=='C1432923':
       L1.append(7)
    elif l1[y]=='C1432924':
       L1.append(8)
    elif l1[y]=='C1432925':
       L1.append(9)
    elif l1[y]=='C1432926':
       L1.append(10)
    elif l1[y]=='C1432927':
       L1.append(11)
    elif l1[y]=='C1432928':
       L1.append(12)
    elif l1[y]=='C1432929':
       L1.append(13)
    elif l1[y]=='C1432931':
       L1.append(14)
    elif l1[y]=='C1432932':
       L1.append(15)
    elif l1[y]=='C1432933':
       L1.append(16)
    elif l1[y]=='C1432934':
       L1.append(17)
    elif l1[y]=='C1432935':
       L1.append(18)
    elif l1[y]=='C1432936':
       L1.append(19)
    elif l1[y]=='C1432937':
       L1.append(20)
    elif l1[y]=='C1432938':
       L1.append(21)
    elif l1[y]=='C1432939':
       L1.append(22)
    elif l1[y]=='C1432940':
       L1.append(23)
    elif l1[y]=='C1432941':
       L1.append(24)
    elif l1[y]=='C1432942':
       L1.append(25)
    elif l1[y]=='C1432943':
       L1.append(26)
    elif l1[y]=='C1432944':
       L1.append(27)
    elif l1[y]=='C1432945':
       L1.append(28)
    elif l1[y]=='C1432946':
       L1.append(29)
    elif l1[y]=='C1432947':
       L1.append(30)
    elif l1[y]=='C1432948':
       L1.append(31)
    elif l1[y]=='C1432949':
       L1.append(32)
    elif l1[y]=='C1432950':
       L1.append(33)
    elif l1[y]=='C1432951':
       L1.append(34)
    elif l1[y]=='C1432952':
       L1.append(35)
    elif l1[y]=='C1432953':
       L1.append(36)
    elif l1[y]=='C1432955':
       L1.append(37)
    elif l1[y]=='C1432956':
       L1.append(38)
    elif l1[y]=='C1432957':
       L1.append(39)
    elif l1[y]=='C1432958':
       L1.append(40)
    elif l1[y]=='C1432959':
       L1.append(41)
    elif l1[y]=='C1471297':
       L1.append(42)
    elif l1[y]=='C1471298':
       L1.append(43)
    elif l1[y]=='C1471299':
       L1.append(44)
    elif l1[y]=='C1471300':
       L1.append(45)
    elif l1[y]=='C1471301':
       L1.append(46)
    elif l1[y]=='C1471302':
       L1.append(47)
    elif l1[y]=='C1471303':
       L1.append(48)
    elif l1[y]=='C1471304':
       L1.append(49)
    elif l1[y]=='C1471305':
       L1.append(50)
    elif l1[y]=='C1471306':
       L1.append(51)
    elif l1[y]=='C1471307':
       L1.append(52)
    elif l1[y]=='C1471308':
       L1.append(53)
    elif l1[y]=='C1471309':
       L1.append(54)
    elif l1[y]=='C1471310':
       L1.append(55)
    elif l1[y]=='C1471311':
       L1.append(56)
    elif l1[y]=='C1471312':
       L1.append(57)
    elif l1[y]=='C1471313':
       L1.append(58)
    elif l1[y]=='C1471314':
       L1.append(59)
    elif l1[y]=='C1471315':
       L1.append(60)
    elif l1[y]=='C1471316':
       L1.append(61)
    elif l1[y]=='C1471317':
       L1.append(62)
    elif l1[y]=='C1471318':
       L1.append(63)
    elif l1[y]=='C1471319':
       L1.append(64)
    elif l1[y]=='C1471320':
       L1.append(65)
    elif l1[y]=='C1471321':
       L1.append(66)
    elif l1[y]=='C1471322':
       L1.append(67)
    elif l1[y]=='C1471323':
       L1.append(68)
    elif l1[y]=='C1471324':
       L1.append(69)
    elif l1[y]=='C1471325':
       L1.append(70)
    elif l1[y]=='C1471326':
       L1.append(71)
    elif l1[y]=='C1471327':
       L1.append(72)
    elif l1[y]=='C1471328':
       L1.append(73)
    elif l1[y]=='C1471329':
       L1.append(74)
    elif l1[y]=='C1471330':
       L1.append(75)
    elif l1[y]=='C1471331':
       L1.append(76)
    elif l1[y]=='C1471332':
       L1.append(77)
    elif l1[y]=='C1471333':
       L1.append(78)
    elif l1[y]=='C1471334':
       L1.append(79)
    elif l1[y]=='C1471335':
       L1.append(80)
    elif l1[y]=='C1471336':
       L1.append(81)
    elif l1[y]=='C1471337':
       L1.append(82)
    elif l1[y]=='C1471338':
       L1.append(83)
    elif l1[y]=='C1471339':
       L1.append(84)
    elif l1[y]=='C1471340':
       L1.append(85)
    elif l1[y]=='C1471341':
       L1.append(86)
    elif l1[y]=='C1471342':
       L1.append(87)
    elif l1[y]=='C1471343':
       L1.append(88)
    elif l1[y]=='C1471344':
       L1.append(89)
    elif l1[y]=='C1471345':
       L1.append(90)
    elif l1[y]=='C1471346':
       L1.append(91)
    elif l1[y]=='C1471347':
       L1.append(92)
    elif l1[y]=='C1471348':
       L1.append(93)
    elif l1[y]=='C1471349':
       L1.append(94)
    elif l1[y]=='C1471350':
       L1.append(95)
    elif l1[y]=='C1471351':
       L1.append(96)
    elif l1[y]=='C1471352':
       L1.append(97)
    elif l1[y]=='C1471353':
       L1.append(98)
    elif l1[y]=='C1471354':
       L1.append(99)
    elif l1[y]=='C1471355':
       L1.append(100)
    elif l1[y]=='C1471356':
       L1.append(101)
    elif l1[y]=='C1513090':
       L1.append(102)
    elif l1[y]=='C1513091':
       L1.append(103)
    elif l1[y]=='C1513092':
       L1.append(104)
    elif l1[y]=='C1513093':
       L1.append(105)
    elif l1[y]=='C1513094':
       L1.append(106)
    elif l1[y]=='C1513095':
       L1.append(107)
    elif l1[y]=='C1513096':
       L1.append(108)
    elif l1[y]=='C1513097':
       L1.append(109)
    elif l1[y]=='C1513098':
       L1.append(110)
    elif l1[y]=='C1513099':
       L1.append(111)
    elif l1[y]=='C1513100':
       L1.append(112)
    elif l1[y]=='C1513101':
       L1.append(113)
    elif l1[y]=='C1513102':
       L1.append(114)
    elif l1[y]=='C1513103':
       L1.append(115)
    elif l1[y]=='C1513104':
       L1.append(116)
    elif l1[y]=='C1513105':
       L1.append(117)
    elif l1[y]=='C1513106':
       L1.append(118)
    elif l1[y]=='C1513107':
       L1.append(119)
    elif l1[y]=='C1513108':
       L1.append(120)
    elif l1[y]=='C1513109':
       L1.append(121)
    elif l1[y]=='C1513110':
       L1.append(122)
    elif l1[y]=='C1513111':
       L1.append(123)
    elif l1[y]=='C1513112':
       L1.append(124)
    elif l1[y]=='C1513113':
       L1.append(125)
    elif l1[y]=='C1513114':
       L1.append(126)
    elif l1[y]=='C1513115':
       L1.append(127)
    elif l1[y]=='C1513116':
       L1.append(128)
    elif l1[y]=='C1513117':
       L1.append(129)
    elif l1[y]=='C1513118':
       L1.append(130)
    elif l1[y]=='C1513119':
       L1.append(131)
    elif l1[y]=='C1513120':
       L1.append(132)
    elif l1[y]=='C1513121':
       L1.append(133)
    elif l1[y]=='C1513122':
       L1.append(134)
    elif l1[y]=='C1513123':
       L1.append(135)
    elif l1[y]=='C1513124':
       L1.append(136)
    elif l1[y]=='C1513126':
       L1.append(137)
    elif l1[y]=='C1513127':
       L1.append(138)
    elif l1[y]=='C1513128':
       L1.append(139)
    elif l1[y]=='C1513129':
       L1.append(140)
    elif l1[y]=='C1513130':
       L1.append(141)
    elif l1[y]=='C1513131':
       L1.append(142)
    elif l1[y]=='C1513132':
       L1.append(143)
    elif l1[y]=='C1513133':
       L1.append(144)
    elif l1[y]=='C1513134':
       L1.append(145)
    elif l1[y]=='C1513135':
       L1.append(146)
    elif l1[y]=='C1513136':
       L1.append(147)
    elif l1[y]=='C1513137':
       L1.append(148)
    elif l1[y]=='C1513138':
       L1.append(149)
    elif l1[y]=='C1513139':
       L1.append(150)
    elif l1[y]=='C1513140':
       L1.append(151)     
    elif l1[y]=='C1513141':
       L1.append(152)
    elif l1[y]=='C1513142':
       L1.append(153)
    elif l1[y]=='C1513143':
       L1.append(154)
    elif l1[y]=='C1513144':
       L1.append(155)
    elif l1[y]=='C1513145':
       L1.append(156)
    elif l1[y]=='C1513146':
       L1.append(157)
    elif l1[y]=='C1513147':
       L1.append(158)
    elif l1[y]=='C1513148':
       L1.append(159)
    elif l1[y]=='C1513149':
       L1.append(160)
    elif l1[y]=='C1513150':
       L1.append(161)
    elif l1[y]=='C1513151':
       L1.append(162)
    elif l1[y]=='C1513152':
       L1.append(163)
    elif l1[y]=='C1513153':
       L1.append(164)
    elif l1[y]=='C1513154':
       L1.append(165)
    elif l1[y]=='C1513155':
       L1.append(166)
    elif l1[y]=='C1513156':
       L1.append(167)
    elif l1[y]=='C1513157':
       L1.append(168)
    elif l1[y]=='C1513158':
       L1.append(169)
    elif l1[y]=='C1513159':
       L1.append(170)
    elif l1[y]=='C1513160':
       L1.append(171)
    elif l1[y]=='C1513161':
       L1.append(172)
    elif l1[y]=='C1513162':
       L1.append(173)
    elif l1[y]=='C1513163':
       L1.append(174)
    elif l1[y]=='C1513164':
       L1.append(175)
    elif l1[y]=='C1513165':
       L1.append(176)
    elif l1[y]=='C1513166':
       L1.append(177)
    elif l1[y]=='C1513167':
       L1.append(178)
    elif l1[y]=='C1513168':
       L1.append(179)
    elif l1[y]=='C1513169':
       L1.append(180)
    elif l1[y]=='C1513170':
       L1.append(181)
    elif l1[y]=='C1513171':
       L1.append(182)
    elif l1[y]=='C1513172':
       L1.append(183)
    elif l1[y]=='C1513173':
       L1.append(184)
    elif l1[y]=='C1513174':
       L1.append(185)
    elif l1[y]=='C1513175':
       L1.append(186)
    elif l1[y]=='C1513176':
       L1.append(187)
    elif l1[y]=='C1513177':
       L1.append(188)
    elif l1[y]=='C1513178':
       L1.append(189)
    elif l1[y]=='C1513179':
       L1.append(190)
    elif l1[y]=='C1513180':
       L1.append(191)
    elif l1[y]=='C1513181':
       L1.append(192)
    elif l1[y]=='C1513182':
       L1.append(193)
    elif l1[y]=='C1513183':
       L1.append(194)
    elif l1[y]=='C1513184':
       L1.append(195)
    elif l1[y]=='C1513185':
       L1.append(196)
    elif l1[y]=='C1513186':
       L1.append(197)
    elif l1[y]=='C1513187':
       L1.append(198)
    elif l1[y]=='C1513188':
       L1.append(199)
    elif l1[y]=='C1513189':
       L1.append(200)
    elif l1[y]=='C1513190':
       L1.append(201)
    elif l1[y]=='C1513191':
       L1.append(202)
    elif l1[y]=='C1513192':
       L1.append(203)
    elif l1[y]=='C1513193':
       L1.append(204)
    elif l1[y]=='C1513194':
       L1.append(205)
    elif l1[y]=='C1513195':
       L1.append(206)
    elif l1[y]=='C1513196':
       L1.append(207)
    elif l1[y]=='C1513197':
       L1.append(208)
    elif l1[y]=='C1513198':
       L1.append(209)
    elif l1[y]=='C1513199':
       L1.append(210)
    elif l1[y]=='C1513200':
       L1.append(211)
    elif l1[y]=='C1513201':
       L1.append(212)
    elif l1[y]=='C1513202':
       L1.append(213)
    elif l1[y]=='C1513203':
       L1.append(214)
    elif l1[y]=='C1513204':
       L1.append(215)
    elif l1[y]=='C1513205':
       L1.append(216)
    elif l1[y]=='C1513206':
       L1.append(217)
    elif l1[y]=='C1513207':
       L1.append(218)
    elif l1[y]=='C1513208':
       L1.append(219)
    elif l1[y]=='C1513209':
       L1.append(220)
    elif l1[y]=='C1513210':
       L1.append(221)
    elif l1[y]=='C1513211':
       L1.append(222)
    elif l1[y]=='C1513212':
       L1.append(223)
    elif l1[y]=='C1513213':
       L1.append(224)
    elif l1[y]=='C1513214':
       L1.append(225)
    elif l1[y]=='C1513215':
       L1.append(226)
    elif l1[y]=='C1513216':
       L1.append(227)
    elif l1[y]=='C1513217':
       L1.append(228)
    elif l1[y]=='C1513218':
       L1.append(229)
    elif l1[y]=='C1513219':
       L1.append(230)
    elif l1[y]=='C1513220':
       L1.append(231)
    elif l1[y]=='C1513221':
       L1.append(232)
    elif l1[y]=='C1513222':
       L1.append(233)
    elif l1[y]=='C1513223':
       L1.append(234)
    elif l1[y]=='C1513224':
       L1.append(235)
    elif l1[y]=='C1513225':
       L1.append(236)
    elif l1[y]=='C1513226':
       L1.append(237)
    elif l1[y]=='C1513227':
       L1.append(238)
    elif l1[y]=='C1513228':
       L1.append(239)
    elif l1[y]=='C1513229':
       L1.append(240)
    elif l1[y]=='C1513230':
       L1.append(241)
    elif l1[y]=='C1513231':
       L1.append(242)
    elif l1[y]=='C1513232':
       L1.append(243)
    elif l1[y]=='C1513233':
       L1.append(244)
    elif l1[y]=='C1513234':
       L1.append(245)
    elif l1[y]=='C1513235':
       L1.append(246)
    elif l1[y]=='C1513236':
       L1.append(247)
    elif l1[y]=='C1513237':
       L1.append(248)
    elif l1[y]=='C1513238':
       L1.append(249)
    elif l1[y]=='C1513239':
       L1.append(250)
    elif l1[y]=='C1513240':
       L1.append(251)
    elif l1[y]=='C1513241':
       L1.append(252)
    elif l1[y]=='C1513242':
       L1.append(253)
    elif l1[y]=='C1513243':
       L1.append(254)
    elif l1[y]=='C1513244':
       L1.append(255)
    elif l1[y]=='C1513245':
       L1.append(256)
    elif l1[y]=='C1513246':
       L1.append(257)
    elif l1[y]=='C1513247':
       L1.append(258)
    elif l1[y]=='C1513248':
       L1.append(259)
    elif l1[y]=='C1513249':
       L1.append(260)
    elif l1[y]=='C1513250':
       L1.append(261)
    elif l1[y]=='C1513251':
       L1.append(262)
    elif l1[y]=='C1513252':
       L1.append(263)
    elif l1[y]=='C1513253':
       L1.append(264)
    elif l1[y]=='C1513254':
       L1.append(265)
    elif l1[y]=='C1513255':
       L1.append(266)
    elif l1[y]=='C1513256':
       L1.append(267)
    elif l1[y]=='C1513257':
       L1.append(268)
    elif l1[y]=='C1513258':
       L1.append(269)
    elif l1[y]=='C1513259':
       L1.append(270)
    elif l1[y]=='C1513260':
       L1.append(271)
    elif l1[y]=='C1513261':
       L1.append(272)
    elif l1[y]=='C1513262':
       L1.append(273)
    elif l1[y]=='C1513263':
       L1.append(274)
    elif l1[y]=='C1513264':
       L1.append(275)
    elif l1[y]=='C1513265':
       L1.append(276)
    elif l1[y]=='C1513266':
       L1.append(277)
    elif l1[y]=='C1513267':
       L1.append(278)
    elif l1[y]=='C1513268':
       L1.append(279)
    elif l1[y]=='C1513269':
       L1.append(280)
    elif l1[y]=='C1513270':
       L1.append(281)
    elif l1[y]=='C1513271':
       L1.append(282)
    elif l1[y]=='C1513272':
       L1.append(283)
    elif l1[y]=='C1513273':
       L1.append(284)
    elif l1[y]=='C1513274':
       L1.append(285)
    elif l1[y]=='C1513275':
       L1.append(286)
    elif l1[y]=='C1513276':
       L1.append(287)
    elif l1[y]=='C1513277':
       L1.append(288)
    elif l1[y]=='C1513278':
       L1.append(289)
    elif l1[y]=='C1513279':
       L1.append(290)
    elif l1[y]=='C1513280':
       L1.append(291)
    elif l1[y]=='C1513281':
       L1.append(292)
    elif l1[y]=='C1513282':
       L1.append(293)
    elif l1[y]=='C1513283':
       L1.append(294)
    elif l1[y]=='C1513285':
       L1.append(295)
    elif l1[y]=='C1513286':
       L1.append(296)
    elif l1[y]=='C1513287':
       L1.append(297)
    elif l1[y]=='C1513288':
       L1.append(298)
    elif l1[y]=='C1513289':
       L1.append(299)
    elif l1[y]=='C1513290':
       L1.append(300)
    elif l1[y]=='C1513291':
       L1.append(301)
    elif l1[y]=='C1618585':
       L1.append(302)
    elif l1[y]=='C1712423':
       L1.append(303)
    elif l1[y]=='C1712424':
       L1.append(304)
    elif l1[y]=='C1712425':
       L1.append(305)
    elif l1[y]=='C1712426':
       L1.append(306)
    elif l1[y]=='C1712427':
       L1.append(307)
    elif l1[y]=='C1712428':
       L1.append(308)
    elif l1[y]=='C1712429':
       L1.append(309)
    elif l1[y]=='C1712430':
       L1.append(310)
    elif l1[y]=='C1712431':
       L1.append(311)
    elif l1[y]=='C1712432':
       L1.append(312)
    elif l1[y]=='C1712433':
       L1.append(313)
    elif l1[y]=='C1712434':
       L1.append(314)
    elif l1[y]=='C1712435':
       L1.append(315)
    elif l1[y]=='C1712437':
       L1.append(316)
    elif l1[y]=='C1712438':
       L1.append(317)
    elif l1[y]=='C1712439':
       L1.append(318)
    elif l1[y]=='C1712440':
       L1.append(319)
    elif l1[y]=='C1712441':
       L1.append(320)
    elif l1[y]=='C1712442':
       L1.append(321)
    elif l1[y]=='C1712443':
       L1.append(322)
    elif l1[y]=='C1765146':
       L1.append(323)
    elif l1[y]=='C1765147':
       L1.append(324)
    elif l1[y]=='C1765148':
       L1.append(325)
    elif l1[y]=='C1765149':
       L1.append(326)
    elif l1[y]=='C1765151':
       L1.append(327)
    elif l1[y]=='C1765152':
       L1.append(328)
    elif l1[y]=='C1765154':
       L1.append(329)
    elif l1[y]=='C1765155':
       L1.append(330)
    elif l1[y]=='C1765156':
       L1.append(331)
    elif l1[y]=='C1765157':
       L1.append(332)
    elif l1[y]=='C1765158':
       L1.append(333)
    elif l1[y]=='C1765159':
       L1.append(334)
    elif l1[y]=='C1765160':
       L1.append(335)
    elif l1[y]=='C1765161':
       L1.append(336)
    elif l1[y]=='C1765162':
       L1.append(337)
    elif l1[y]=='C1765163':
       L1.append(338)
    elif l1[y]=='C1765164':
       L1.append(339)
    elif l1[y]=='C1765165':
       L1.append(340)
    elif l1[y]=='C1765166':
       L1.append(341)
    elif l1[y]=='C1765167':
       L1.append(342)
    elif l1[y]=='C1768804':
       L1.append(343)
    elif l1[y]=='C1768806':
       L1.append(344)
    elif l1[y]=='C1768808':
       L1.append(345)
    elif l1[y]=='C1768809':
       L1.append(346)
    elif l1[y]=='C1768811':
       L1.append(347)
    elif l1[y]=='C1768813':
       L1.append(348)
    elif l1[y]=='C1768815':
       L1.append(349)
    elif l1[y]=='C1768816':
       L1.append(350)
    elif l1[y]=='C1768818':
       L1.append(351)
    elif l1[y]=='C1768820':
       L1.append(352)
    elif l1[y]=='C1768822':
       L1.append(353)
    elif l1[y]=='C1768823':
       L1.append(354)
    elif l1[y]=='C1768825':
       L1.append(355)
    elif l1[y]=='C1768826':
       L1.append(356)
    elif l1[y]=='C1768827':
       L1.append(357)
    elif l1[y]=='C1768828':
       L1.append(358)
    elif l1[y]=='C1768829':
       L1.append(359)
    elif l1[y]=='C1768830':
       L1.append(360)
    elif l1[y]=='C1768831':
       L1.append(361)
    elif l1[y]=='C1768832':
       L1.append(362)
    elif l1[y]=='C1784063':
       L1.append(363)
    elif l1[y]=='C1784064':
       L1.append(364)
    elif l1[y]=='C1784065':
       L1.append(365)
    elif l1[y]=='C1784066':
       L1.append(366)
    elif l1[y]=='C1784067':
       L1.append(367)
    elif l1[y]=='C1784068':
       L1.append(368)
    elif l1[y]=='C1784069':
       L1.append(369)
    elif l1[y]=='C1784070':
       L1.append(370)
    elif l1[y]=='C1784071':
       L1.append(371)
    elif l1[y]=='C1784072':
       L1.append(372)
    elif l1[y]=='C1784073':
       L1.append(373)
    elif l1[y]=='C1784074':
       L1.append(374)
    elif l1[y]=='C1784075':
       L1.append(375)
    elif l1[y]=='C1784076':
       L1.append(376)
    elif l1[y]=='C1784077':
       L1.append(377)
    elif l1[y]=='C1784078':
       L1.append(378)
    elif l1[y]=='C1784079':
       L1.append(379)
    elif l1[y]=='C1784080':
       L1.append(380)
    elif l1[y]=='C1784081':
       L1.append(381)
    elif l1[y]=='C1784082':
       L1.append(382)
    elif l1[y]=='C1793522':
       L1.append(383)
    elif l1[y]=='C1793523':
       L1.append(384)
    elif l1[y]=='C1793524':
       L1.append(385)
    elif l1[y]=='C1793525':
       L1.append(386)
    elif l1[y]=='C1793526':
       L1.append(387)
    elif l1[y]=='C1793527':
       L1.append(388)
    elif l1[y]=='C1793528':
       L1.append(389)
    elif l1[y]=='C1793530':
       L1.append(390)
    elif l1[y]=='C1793531':
       L1.append(391)
    elif l1[y]=='C1793532':
       L1.append(392)
    elif l1[y]=='C1793533':
       L1.append(393)
    elif l1[y]=='C1793534':
       L1.append(394)
    elif l1[y]=='C1793535':
       L1.append(395)
    elif l1[y]=='C1793536':
       L1.append(396)
    elif l1[y]=='C1793537':
       L1.append(397)
    elif l1[y]=='C1793538':
       L1.append(398)
    elif l1[y]=='C1793539':
       L1.append(399)
    elif l1[y]=='C1793540':
       L1.append(400)
    elif l1[y]=='C1793542':
       L1.append(401)
    elif l1[y]=='C1793543':
       L1.append(402)
    elif l1[y]=='C1802309':
       L1.append(403)
    elif l1[y]=='C1802310':
       L1.append(404)
    elif l1[y]=='C1802311':
       L1.append(405)
    elif l1[y]=='C1802312':
       L1.append(406)
    elif l1[y]=='C1802313':
       L1.append(407)
    elif l1[y]=='C1802314':
       L1.append(408)
    elif l1[y]=='C1802315':
       L1.append(409)
    elif l1[y]=='C1802317':
       L1.append(410)
    elif l1[y]=='C1802318':
       L1.append(411)
    elif l1[y]=='C1802319':
       L1.append(412)
    elif l1[y]=='C1802320':
       L1.append(413)
    elif l1[y]=='C1802321':
       L1.append(414)
    elif l1[y]=='C1802322':
       L1.append(415)
    elif l1[y]=='C1802323':
       L1.append(416)
    elif l1[y]=='C1802324':
       L1.append(417)
    elif l1[y]=='C1802325':
       L1.append(418)
    elif l1[y]=='C1802326':
       L1.append(419)
    elif l1[y]=='C1802327':
       L1.append(420)
    elif l1[y]=='C1802328':
       L1.append(421)
    elif l1[y]=='C1802329':
       L1.append(422)
    elif l1[y]=='C1804587':
       L1.append(423)
    elif l1[y]=='C1804588':
       L1.append(424)
    elif l1[y]=='C1804589':
       L1.append(425)
    elif l1[y]=='C1804590':
       L1.append(426)
    elif l1[y]=='C1804591':
       L1.append(427)
    elif l1[y]=='C1804592':
       L1.append(428)
    elif l1[y]=='C1804593':
       L1.append(429)
    elif l1[y]=='C1804594':
       L1.append(430)
    elif l1[y]=='C1804595':
       L1.append(431)
    elif l1[y]=='C1804596':
       L1.append(432)
    elif l1[y]=='C1804597':
       L1.append(433)
    elif l1[y]=='C1804598':
       L1.append(434)
    elif l1[y]=='C1804599':
       L1.append(435)
    elif l1[y]=='C1804601':
       L1.append(436)
    elif l1[y]=='C1804602':
       L1.append(437)
    elif l1[y]=='C1804603':
       L1.append(438)
    elif l1[y]=='C1804604':
       L1.append(439)
    elif l1[y]=='C1804605':
       L1.append(440)
    elif l1[y]=='C1804606':
       L1.append(441)
    elif l1[y]=='C1804607':
       L1.append(442)
    elif l1[y]=='C1825882':
       L1.append(443)
    elif l1[y]=='C1825883':
       L1.append(444)
    elif l1[y]=='C1825885':
       L1.append(445)
    elif l1[y]=='C1825886':
       L1.append(446)
    elif l1[y]=='C1825887':
       L1.append(447)
    elif l1[y]=='C1825888':
       L1.append(448)
    elif l1[y]=='C1825889':
       L1.append(449)
    elif l1[y]=='C1825890':
       L1.append(450)
    elif l1[y]=='C1825891':
       L1.append(451)
    elif l1[y]=='C1825892':
       L1.append(452)
    elif l1[y]=='C1825893':
       L1.append(453)
    elif l1[y]=='C1825894':
       L1.append(454)
    elif l1[y]=='C1825895':
       L1.append(455)
    elif l1[y]=='C1825897':
       L1.append(456)
    elif l1[y]=='C1825898':
       L1.append(457)
    elif l1[y]=='C1825900':
       L1.append(458)
    elif l1[y]=='C1825901':
       L1.append(459)
    elif l1[y]=='C1825902':
       L1.append(460)
    elif l1[y]=='C1825903':
       L1.append(461)
    elif l1[y]=='C1825904':
       L1.append(462)
    elif l1[y]=='C1847303':
       L1.append(463)
    elif l1[y]=='C1847304':
       L1.append(464)
    elif l1[y]=='C1847305':
       L1.append(465)
    elif l1[y]=='C1847306':
       L1.append(466)
    elif l1[y]=='C1847307':
       L1.append(467)
    elif l1[y]=='C1847309':
       L1.append(468)
    elif l1[y]=='C1847310':
       L1.append(469)
    elif l1[y]=='C1847311':
       L1.append(470)
    elif l1[y]=='C1847312':
       L1.append(471)
    elif l1[y]=='C1847313':
       L1.append(472)
    elif l1[y]=='C1847314':
       L1.append(473)
    elif l1[y]=='C1847315':
       L1.append(474)
    elif l1[y]=='C1847316':
       L1.append(475)
    elif l1[y]=='C1847317':
       L1.append(476)
    elif l1[y]=='C1847319':
       L1.append(477)
    elif l1[y]=='C1847320':
       L1.append(478)
    elif l1[y]=='C1847321':
       L1.append(479)
    elif l1[y]=='C1847322':
       L1.append(480)
    elif l1[y]=='C1847323':
       L1.append(481)
    elif l1[y]=='C1847324':
       L1.append(482)
    elif l1[y]=='C2700874':
       L1.append(483)
    elif l1[y]=='C2700877':
       L1.append(484)
    elif l1[y]=='C2700880':
       L1.append(485)
    elif l1[y]=='C2700883':
       L1.append(486)
    elif l1[y]=='C2700886':
       L1.append(487)
    elif l1[y]=='C2700889':
       L1.append(488)
    elif l1[y]=='C2700892':
       L1.append(489)
    elif l1[y]=='C2700895':
       L1.append(490)
    elif l1[y]=='C2700898':
       L1.append(491)
    elif l1[y]=='C2700901':
       L1.append(492)
    elif l1[y]=='C5115391':
       L1.append(493)
    elif l1[y]=='C5115394':
       L1.append(494)
    elif l1[y]=='C5115397':
       L1.append(495)
    elif l1[y]=='C5115400':
       L1.append(496)
    elif l1[y]=='C5115403':
       L1.append(497)
    elif l1[y]=='C5115406':
       L1.append(498)
    elif l1[y]=='C5115409':
       L1.append(499)
    elif l1[y]=='C5115412':
       L1.append(500)   
    elif  l1[y]=='C1857307':
        L1.append(501)
    elif  l1[y]=='C1857308':
        L1.append(502)
    elif  l1[y]=='C1857309':
        L1.append(503)
    elif  l1[y]=='C1857311':
        L1.append(504)
    elif  l1[y]=='C1857312':
        L1.append(505)
    elif  l1[y]=='C1857313':
        L1.append(506)
    elif  l1[y]=='C1857314':
        L1.append(507)
    elif  l1[y]=='C1857315':
        L1.append(508)
    elif  l1[y]=='C1857316':
        L1.append(509)
    elif  l1[y]=='C1857317':
        L1.append(510)
    elif  l1[y]=='C1857318':
        L1.append(511)
    elif  l1[y]=='C1857319':
        L1.append(512)
    elif  l1[y]=='C1857320':
        L1.append(513)
    elif  l1[y]=='C1857321':
        L1.append(514)
    elif  l1[y]=='C1857322':
        L1.append(515)
    elif  l1[y]=='C1857323':
        L1.append(516)
    elif  l1[y]=='C1857324':
        L1.append(517)
    elif  l1[y]=='C1857325':
        L1.append(518)
    elif  l1[y]=='C1857326':
        L1.append(519)
    elif  l1[y]=='C1857327':
        L1.append(520)
    elif  l1[y]=='C1857328':
        L1.append(521)
    elif  l1[y]=='C1857329':
        L1.append(522)
    elif  l1[y]=='C1857330':
        L1.append(523)
    elif  l1[y]=='C1857331':
        L1.append(524)
    elif  l1[y]=='C1857332':
        L1.append(525)
    elif  l1[y]=='C1857334':
        L1.append(526)
    elif  l1[y]=='C1857335':
        L1.append(527)
    elif  l1[y]=='C1857336':
        L1.append(528)
    elif  l1[y]=='C1857337':
        L1.append(529)
    elif  l1[y]=='C1857338':
        L1.append(530)
    elif  l1[y]=='C1857339':
        L1.append(531)
    elif  l1[y]=='C1857341':
        L1.append(532)
    elif  l1[y]=='C1857342':
        L1.append(533)
    elif  l1[y]=='C1857344':
        L1.append(534)
    elif  l1[y]=='C1857345':
        L1.append(535)
    elif  l1[y]=='C1857346':
        L1.append(536)
    elif  l1[y]=='C1857347':
        L1.append(537)
    elif  l1[y]=='C1857348':
        L1.append(538)
    elif  l1[y]=='C1857349':
        L1.append(539)
    elif  l1[y]=='C1857351':
        L1.append(540)
    elif  l1[y]=='C1895132':
        L1.append(541)
    elif  l1[y]=='C1895133':
        L1.append(542)
    elif  l1[y]=='C1895134':
        L1.append(543)
    elif  l1[y]=='C1895135':
        L1.append(544)
    elif  l1[y]=='C1895136':
        L1.append(545)
    elif  l1[y]=='C1895137':
        L1.append(546)
    elif  l1[y]=='C1895138':
        L1.append(547)
    elif  l1[y]=='C1895139':
        L1.append(548)
    elif  l1[y]=='C1895140':
        L1.append(549)
    elif  l1[y]=='C1895141':
        L1.append(550)
    elif  l1[y]=='C1895142':
        L1.append(551)
    elif  l1[y]=='C1895143':
        L1.append(552)
    elif  l1[y]=='C1895144':
        L1.append(553)
    elif  l1[y]=='C1895145':
        L1.append(554)
    elif  l1[y]=='C1895146':
        L1.append(555)
    elif  l1[y]=='C1895147':
        L1.append(556)
    elif  l1[y]=='C1895148':
        L1.append(557)
    elif  l1[y]=='C1895149':
        L1.append(558)
    elif  l1[y]=='C1895150':
        L1.append(559)
    elif  l1[y]=='C1895151':
        L1.append(560)
    elif  l1[y]=='C2058381':
        L1.append(561)
    elif  l1[y]=='C2058384':
        L1.append(562)
    elif  l1[y]=='C2058387':
        L1.append(563)
    elif  l1[y]=='C2058390':
        L1.append(564)
    elif  l1[y]=='C2058393':
        L1.append(565)
    elif  l1[y]=='C2058396':
        L1.append(566)
    elif  l1[y]=='C2058399':
        L1.append(567)
    elif  l1[y]=='C2058402':
        L1.append(568)
    elif  l1[y]=='C4739615':
        L1.append(569)
    elif  l1[y]=='C2058408':
        L1.append(570)
    elif  l1[y]=='C2058411':
        L1.append(571)
    elif  l1[y]=='C2058414':
        L1.append(572)
    elif  l1[y]=='C2058417':
        L1.append(573)
    elif  l1[y]=='C2058420':
        L1.append(574)
    elif  l1[y]=='C2058423':
        L1.append(575)
    elif  l1[y]=='C2058426':
        L1.append(576)
    elif  l1[y]=='C2058429':
        L1.append(577)
    elif  l1[y]=='C2058432':
        L1.append(578)
    elif  l1[y]=='C2058435': 
        L1.append(579)
    elif  l1[y]=='C2058438': 
        L1.append(580)
    elif  l1[y]=='C2385888': 
        L1.append(581)
    elif  l1[y]=='C2385891': 
        L1.append(582)
    elif  l1[y]=='C2385894': 
        L1.append(583)
    elif  l1[y]=='C2385897': 
        L1.append(584)
    elif  l1[y]=='C2385900': 
        L1.append(585)
    elif  l1[y]=='C2385903':
        L1.append(586)
    elif  l1[y]=='C2385906': 
        L1.append(587)
    elif  l1[y]=='C2385909': L1.append(588)
    elif  l1[y]=='C2385912': L1.append(589)
    elif  l1[y]=='C2385915': L1.append(590)
    elif  l1[y]=='C2461964': L1.append(591)
    elif  l1[y]=='C2461967': L1.append(592)
    elif  l1[y]=='C2461973': L1.append(593)
    elif  l1[y]=='C2461976': L1.append(594)
    elif  l1[y]=='C2461979': L1.append(595)
    elif  l1[y]=='C2461982': L1.append(596)
    elif  l1[y]=='C2461985': L1.append(597)
    elif  l1[y]=='C2461988': L1.append(598)
    elif  l1[y]=='C2461991': L1.append(599)
    elif  l1[y]=='C2461994': L1.append(600)
    elif  l1[y]=='C2461997': L1.append(601)
    elif  l1[y]=='C2462000': L1.append(602)
    elif  l1[y]=='C2462003': L1.append(603)
    elif  l1[y]=='C2462006': L1.append(604)
    elif  l1[y]=='C2462009': L1.append(605)
    elif  l1[y]=='C2462012': L1.append(606)
    elif  l1[y]=='C2462015': L1.append(607)
    elif  l1[y]=='C2462018': L1.append(608)
    elif  l1[y]=='C2462021': L1.append(609)
    elif  l1[y]=='C2462024': L1.append(610)
    elif  l1[y]=='C2510793': L1.append(611)
    elif  l1[y]=='C2510796': L1.append(612)
    elif  l1[y]=='C2510799': L1.append(613)
    elif  l1[y]=='C2510802': L1.append(614)
    elif  l1[y]=='C2510805': L1.append(615)
    elif  l1[y]=='C2510808': L1.append(616)
    elif  l1[y]=='C2510811': L1.append(617)
    elif  l1[y]=='C2510814': L1.append(618)
    elif  l1[y]=='C2510817': L1.append(619)
    elif  l1[y]=='C2510820': L1.append(620)
    elif  l1[y]=='C2548818': L1.append(621)
    elif  l1[y]=='C2548824': L1.append(622)
    elif  l1[y]=='C2548827': L1.append(623)
    elif  l1[y]=='C2548830': L1.append(624)
    elif  l1[y]=='C2548833': L1.append(625)
    elif  l1[y]=='C2548836': L1.append(626)
    elif  l1[y]=='C2548839': L1.append(627)
    elif  l1[y]=='C2548842': L1.append(628)
    elif  l1[y]=='C2548845': L1.append(629)
    elif  l1[y]=='C2548848': L1.append(630)
    elif  l1[y]=='C2549013': L1.append(631)
    elif  l1[y]=='C2585520': L1.append(632)
    elif  l1[y]=='C2585523': L1.append(633)
    elif  l1[y]=='C2585526': L1.append(634)
    elif  l1[y]=='C2585532': L1.append(635)
    elif  l1[y]=='C2585535': L1.append(636)
    elif  l1[y]=='C2597190': L1.append(637)
    elif  l1[y]=='C2597193': L1.append(638)
    elif  l1[y]=='C2597196': L1.append(639)
    elif  l1[y]=='C2597199': L1.append(640)
    elif  l1[y]=='C2597202': L1.append(641)
    elif  l1[y]=='C2597205': L1.append(642)
    elif  l1[y]=='C2597208': L1.append(643)
    elif  l1[y]=='C2597211': L1.append(644)
    elif  l1[y]=='C2597214': L1.append(645)
    elif  l1[y]=='C2597217': L1.append(646)
    elif  l1[y]=='C2611348': L1.append(647)
    elif  l1[y]=='C2611351': L1.append(648)
    elif  l1[y]=='C2611354': L1.append(649)
    elif  l1[y]=='C2611357': L1.append(650)
    elif  l1[y]=='C2643952': L1.append(651)
    elif  l1[y]=='C2643955': L1.append(652)
    elif  l1[y]=='C2643958': L1.append(653)
    elif  l1[y]=='C2643961': L1.append(654)
    elif  l1[y]=='C2643964': L1.append(655)
    elif  l1[y]=='C2643967': L1.append(656)
    elif  l1[y]=='C2643970': L1.append(657)
    elif  l1[y]=='C2643973': L1.append(658)
    elif  l1[y]=='C2643976': L1.append(659)
    elif  l1[y]=='C2643979': L1.append(660)
    elif  l1[y]=='C2653767': L1.append(661)
    elif  l1[y]=='C2653770': L1.append(662)
    elif  l1[y]=='C2653773': L1.append(663)
    elif  l1[y]=='C2653776': L1.append(664)
    elif  l1[y]=='C2653779': L1.append(665)
    elif  l1[y]=='C2653782': L1.append(666)
    elif  l1[y]=='C2653785': L1.append(667)
    elif  l1[y]=='C2653788': L1.append(668)
    elif  l1[y]=='C2653791': L1.append(669)
    elif  l1[y]=='C2653977': L1.append(670)
    elif  l1[y]=='C2678603': L1.append(671)
    elif  l1[y]=='C2678606': L1.append(672)
    elif  l1[y]=='C2678609': L1.append(673)
    elif  l1[y]=='C2678612': L1.append(674)
    elif  l1[y]=='C2678615': L1.append(675)
    elif  l1[y]=='C2678618': L1.append(676)
    elif  l1[y]=='C2678621': L1.append(677)
    elif  l1[y]=='C2678624': L1.append(678)
    elif  l1[y]=='C2678627': L1.append(679)
    elif  l1[y]=='C2678633': L1.append(680)
    elif  l1[y]=='C2712402': L1.append(681)
    elif  l1[y]=='C2712405': L1.append(682)
    elif  l1[y]=='C2712408': L1.append(683)
    elif  l1[y]=='C2712411': L1.append(684)
    elif  l1[y]=='C4739618': L1.append(685)
    elif  l1[y]=='C2734850': L1.append(686)
    elif  l1[y]=='C2734853': L1.append(687)
    elif  l1[y]=='C2734856': L1.append(688)
    elif  l1[y]=='C2734859': L1.append(689)
    elif  l1[y]=='C2734862': L1.append(690)
    elif  l1[y]=='C2734865': L1.append(691)
    elif  l1[y]=='C2734868': L1.append(692)
    elif  l1[y]=='C2734871': L1.append(693)
    elif  l1[y]=='C2734874': L1.append(694)
    elif  l1[y]=='C2734877': L1.append(695)
    elif  l1[y]=='C2765246': L1.append(696)
    elif  l1[y]=='C2765249': L1.append(697)
    elif  l1[y]=='C2765252': L1.append(698)
    elif  l1[y]=='C2765255': L1.append(699)
    elif  l1[y]=='C2765258': L1.append(700)
    elif  l1[y]=='C2765261': L1.append(701)
    elif  l1[y]=='C2765264': L1.append(702)
    elif  l1[y]=='C2765267': L1.append(703)
    elif  l1[y]=='C2765273': L1.append(704)
    elif  l1[y]=='C2765276': L1.append(705)
    elif  l1[y]=='C2803142': L1.append(706)
    elif  l1[y]=='C2803145': L1.append(707)
    elif  l1[y]=='C2803148': L1.append(708)
    elif  l1[y]=='C2803151': L1.append(709)
    elif  l1[y]=='C2803154': L1.append(710)
    elif  l1[y]=='C2803157': L1.append(711)
    elif  l1[y]=='C2803160': L1.append(712)
    elif  l1[y]=='C2803163': L1.append(713)
    elif  l1[y]=='C2803166': L1.append(714)
    elif  l1[y]=='C2803169': L1.append(715)
    elif  l1[y]=='C2816080': L1.append(716)
    elif  l1[y]=='C2820479': L1.append(717)
    elif  l1[y]=='C2820482': L1.append(718)
    elif  l1[y]=='C2820485': L1.append(719)
    elif  l1[y]=='C2820488': L1.append(720)
    elif  l1[y]=='C2820491': L1.append(721)
    elif  l1[y]=='C2820494': L1.append(722)
    elif  l1[y]=='C2820497': L1.append(723)
    elif  l1[y]=='C2820500': L1.append(724)
    elif  l1[y]=='C2820503': L1.append(725)
    elif  l1[y]=='C2820506': L1.append(726)
    elif  l1[y]=='C2820509': L1.append(727)
    elif  l1[y]=='C2820512': L1.append(728)
    elif  l1[y]=='C2820515': L1.append(729)
    elif  l1[y]=='C2820518': L1.append(730)
    elif  l1[y]=='C2820521': L1.append(731)
    elif  l1[y]=='C2820524': L1.append(732)
    elif  l1[y]=='C2820527': L1.append(733)
    elif  l1[y]=='C2820530': L1.append(734)
    elif  l1[y]=='C2820533': L1.append(735)
    elif  l1[y]=='C2820536': L1.append(736)
    elif  l1[y]=='C2978381': L1.append(737)
    elif  l1[y]=='C2978384': L1.append(738)
    elif  l1[y]=='C2978387': L1.append(739)
    elif  l1[y]=='C2978390': L1.append(740)
    elif  l1[y]=='C2978393': L1.append(741)
    elif  l1[y]=='C2978396': L1.append(742)
    elif  l1[y]=='C2978399': L1.append(743)
    elif  l1[y]=='C2978402': L1.append(744)
    elif  l1[y]=='C2978405': L1.append(745)
    elif  l1[y]=='C2978408': L1.append(746)
    elif  l1[y]=='C3022008': L1.append(747)
    elif  l1[y]=='C3022011': L1.append(748)
    elif  l1[y]=='C3022014': L1.append(749)
    elif  l1[y]=='C3022020': L1.append(750)
    elif  l1[y]=='C3022023': L1.append(751)
    elif  l1[y]=='C3022026': L1.append(752)
    elif  l1[y]=='C3022029': L1.append(753)
    elif  l1[y]=='C3022032': L1.append(754)
    elif  l1[y]=='C3022035': L1.append(755)
    elif  l1[y]=='C3022038': L1.append(756)
    elif  l1[y]=='C3141481': L1.append(757)
    elif  l1[y]=='C3141484': L1.append(758)
    elif  l1[y]=='C3141487': L1.append(759)
    elif  l1[y]=='C3141490': L1.append(760)
    elif  l1[y]=='C3141493': L1.append(761)
    elif  l1[y]=='C3141496': L1.append(762)
    elif  l1[y]=='C3141499': L1.append(763)
    elif  l1[y]=='C3141502': L1.append(764)
    elif  l1[y]=='C3141505': L1.append(765)
    elif  l1[y]=='C3141508': L1.append(766)
    elif  l1[y]=='C3270735': L1.append(767)
    elif  l1[y]=='C3270738': L1.append(768)
    elif  l1[y]=='C3270741': L1.append(769)
    elif  l1[y]=='C3270744': L1.append(770)
    elif l1[y]==53420103236: L1.append(771)
    elif l1[y]==53420103237: L1.append(772)
    elif l1[y]==53420103238: L1.append(773)
    elif l1[y]==53420103239: L1.append(774)
    elif l1[y]==53420103240: L1.append(775)
    elif l1[y]==53420103241: L1.append(776)
    elif l1[y]==53420103242: L1.append(777)
    elif l1[y]==53420103243: L1.append(778)
    elif l1[y]==53420103244: L1.append(779)
    elif l1[y]==53420103245: L1.append(780)
    elif  l1[y]=='C3350679': L1.append(781)
    elif  l1[y]=='C3350682': L1.append(782)
    elif  l1[y]=='C3350685': L1.append(783)
    elif  l1[y]=='C3350688': L1.append(784)
    elif  l1[y]=='C3350691': L1.append(785)
    elif  l1[y]=='C3350694': L1.append(786)
    elif  l1[y]=='C3350697': L1.append(787)
    elif  l1[y]=='C3350700': L1.append(788)
    elif  l1[y]=='C3350703': L1.append(789)
    elif  l1[y]=='C3350706': L1.append(790)
    elif  l1[y]=='C3463380': L1.append(791)
    elif  l1[y]=='C3463383': L1.append(792)
    elif  l1[y]=='C3463386': L1.append(793)
    elif  l1[y]=='C3463389': L1.append(794)
    elif  l1[y]=='C3463392': L1.append(795)
    elif  l1[y]=='C3489065': L1.append(796)
    elif  l1[y]=='C3489068': L1.append(797)
    elif  l1[y]=='C3489071': L1.append(798)
    elif  l1[y]=='C3489074': L1.append(799)
    elif  l1[y]=='C3490380': L1.append(800)
    elif  l1[y]=='C3578837': L1.append(801)
    elif  l1[y]=='C3578840': L1.append(802)
    elif  l1[y]=='C3578843': L1.append(803)
    elif  l1[y]=='C3578846': L1.append(804)
    elif  l1[y]=='C3578849': L1.append(805)
    elif  l1[y]=='C3599509': L1.append(806)
    elif  l1[y]=='C3599512': L1.append(807)
    elif  l1[y]=='C3599515': L1.append(808)
    elif  l1[y]=='C3599518': L1.append(809)
    elif  l1[y]=='C3599521': L1.append(810)
    elif  l1[y]=='C3692781': L1.append(811)
    elif  l1[y]=='C3692784': L1.append(812)
    elif  l1[y]=='C3692787': L1.append(813)
    elif  l1[y]=='C3692790': L1.append(814)
    elif  l1[y]=='C3692793': L1.append(815)
    elif  l1[y]=='C3745413': L1.append(816)
    elif  l1[y]=='C3745416': L1.append(817)
    elif  l1[y]=='C3745419': L1.append(818)
    elif  l1[y]=='C3745425': L1.append(819)
    elif  l1[y]=='C3745428': L1.append(820)
    elif  l1[y]=='C3799615': L1.append(821)
    elif  l1[y]=='C3799618': L1.append(822)
    elif  l1[y]=='C3799621': L1.append(823)
    elif  l1[y]=='C3799624': L1.append(824)
    elif  l1[y]=='C3799998': L1.append(825)
    elif  l1[y]=='C3878043': L1.append(826)
    elif  l1[y]=='C3878046': L1.append(827)
    elif  l1[y]=='C3878049': L1.append(828)
    elif  l1[y]=='C3878052': L1.append(829)
    elif  l1[y]=='C3878055': L1.append(830)
    elif  l1[y]=='C3942400': L1.append(831)
    elif  l1[y]=='C3942403': L1.append(832)
    elif  l1[y]=='C3942406': L1.append(833)
    elif  l1[y]=='C3942409': L1.append(834)
    elif  l1[y]=='C3942412': L1.append(835)
    elif  l1[y]=='C3945572': L1.append(836)
    elif  l1[y]=='C4031503': L1.append(837)
    elif  l1[y]=='C4031506': L1.append(838)
    elif  l1[y]=='C4031509': L1.append(839)
    elif  l1[y]=='C4031512': L1.append(840)
    elif  l1[y]=='C4031515': L1.append(841)
    elif  l1[y]=='C4113144': L1.append(842)
    elif  l1[y]=='C4113147': L1.append(843)
    elif  l1[y]=='C4113150': L1.append(844)
    elif  l1[y]=='C4113153': L1.append(845)
    elif  l1[y]=='C4113156': L1.append(846)
    elif  l1[y]=='C4129542': L1.append(847)
    elif  l1[y]=='C4189296': L1.append(848)
    elif  l1[y]=='C4189299': L1.append(849)
    elif  l1[y]=='C4189302': L1.append(850)
    elif  l1[y]=='C4189305': L1.append(851)
    elif  l1[y]=='C4189308': L1.append(852)
    elif  l1[y]=='C4251677': L1.append(853)
    elif  l1[y]=='C4257541': L1.append(854)
    elif  l1[y]=='C4257544': L1.append(855)
    elif  l1[y]=='C4257547': L1.append(856)
    elif  l1[y]=='C4257550': L1.append(857)
    elif  l1[y]=='C4257553': L1.append(858)
    elif  l1[y]=='C4273197': L1.append(859)
    elif  l1[y]=='C4327180': L1.append(860)
    elif  l1[y]=='C4332698': L1.append(861)
    elif  l1[y]=='C4332701': L1.append(862)
    elif  l1[y]=='C4332704': L1.append(863)
    elif  l1[y]=='C4332707': L1.append(864)
    elif  l1[y]=='C4332710': L1.append(865)
    elif  l1[y]=='C4347525': L1.append(866)
    elif  l1[y]=='C4347528': L1.append(867)
    elif  l1[y]=='C4347531': L1.append(868)
    elif  l1[y]=='C4347534': L1.append(869)
    elif  l1[y]=='C4347537': L1.append(870)
    elif l1[y]==53420101310: L1.append(871)
    elif l1[y]==53420101311: L1.append(872)
    elif l1[y]==53420101312: L1.append(873)
    elif l1[y]==53420101313: L1.append(874)
    elif l1[y]==53420101314: L1.append(875)
    elif  l1[y]=='C4511352': L1.append(876)
    elif  l1[y]=='C4511355': L1.append(877)
    elif  l1[y]=='C4511358': L1.append(878)
    elif  l1[y]=='C4511361': L1.append(879)
    elif  l1[y]=='C4511364': L1.append(880)
    elif  l1[y]=='C4549953': L1.append(881)
    elif  l1[y]=='C4549956': L1.append(882)
    elif  l1[y]=='C4549959': L1.append(883)
    elif  l1[y]=='C4549962': L1.append(884)
    elif  l1[y]=='C4549965': L1.append(885)
    elif  l1[y]=='C4613602': L1.append(886)
    elif  l1[y]=='C4613605': L1.append(887)
    elif  l1[y]=='C4613608': L1.append(888)
    elif  l1[y]=='C4613611': L1.append(889)
    elif  l1[y]=='C4613614': L1.append(890)
    elif  l1[y]=='C4613617': L1.append(891)
    elif  l1[y]=='C4613620': L1.append(892)
    elif  l1[y]=='C4613623': L1.append(893)
    elif  l1[y]=='C4613626': L1.append(894)
    elif  l1[y]=='C4613629': L1.append(895)
    elif  l1[y]=='C4613632': L1.append(896)
    elif  l1[y]=='C4613635': L1.append(897)
    elif  l1[y]=='C4613638': L1.append(898)
    elif  l1[y]=='C4613641': L1.append(899)
    elif  l1[y]=='C4613647': L1.append(900)
    elif  l1[y]=='C4613650': L1.append(901)
    elif  l1[y]=='C4613653': L1.append(902)
    elif  l1[y]=='C4613656': L1.append(903)
    elif  l1[y]=='C4613662': L1.append(904)
    elif  l1[y]=='C4613665': L1.append(905)
    elif  l1[y]=='C4627412': L1.append(906)
    elif  l1[y]=='C4627415': L1.append(907)
    elif  l1[y]=='C4627418': L1.append(908)
    elif  l1[y]=='C4627421': L1.append(909)
    elif  l1[y]=='C4627424': L1.append(910)
    elif  l1[y]=='C4627427': L1.append(911)
    elif  l1[y]=='C4627430': L1.append(912)
    elif  l1[y]=='C4627433': L1.append(913)
    elif  l1[y]=='C4627436': L1.append(914)
    elif  l1[y]=='C4627439': L1.append(915)
    elif  l1[y]=='C4627442': L1.append(916)
    elif  l1[y]=='C4627445': L1.append(917)
    elif  l1[y]=='C4627448': L1.append(918)
    elif  l1[y]=='C4627451': L1.append(919)
    elif  l1[y]=='C4627454': L1.append(920)
    elif  l1[y]=='C4627457': L1.append(921)
    elif  l1[y]=='C4627460': L1.append(922)
    elif  l1[y]=='C4627463': L1.append(923)
    elif  l1[y]=='C4627466': L1.append(924)
    elif  l1[y]=='C4627469': L1.append(925)
    elif  l1[y]=='C4710087': L1.append(926)
    elif  l1[y]=='C4710090': L1.append(927)
    elif  l1[y]=='C4710093': L1.append(928)
    elif  l1[y]=='C4710096': L1.append(929)
    elif  l1[y]=='C4710099': L1.append(930)
    elif  l1[y]=='C4722444': L1.append(931)
    elif  l1[y]=='C4722447': L1.append(932)
    elif  l1[y]=='C4722450': L1.append(933)
    elif  l1[y]=='C4722453': L1.append(934)
    elif  l1[y]=='C4722456': L1.append(935)
    elif  l1[y]=='C4722459': L1.append(936)
    elif  l1[y]=='C4722462': L1.append(937)
    elif  l1[y]=='C4722465': L1.append(938)
    elif  l1[y]=='C4722468': L1.append(939)
    elif  l1[y]=='C4722471': L1.append(940)
    elif  l1[y]=='C4748870': L1.append(941)
    elif  l1[y]=='C4748873': L1.append(942)
    elif  l1[y]=='C4748876': L1.append(943)
    elif  l1[y]=='C4748882': L1.append(944)
    elif  l1[y]=='C4748885': L1.append(945)
    elif  l1[y]=='C4794445': L1.append(946)
    elif  l1[y]=='C4794448': L1.append(947)
    elif  l1[y]=='C4794451': L1.append(948)
    elif  l1[y]=='C4794454': L1.append(949)
    elif  l1[y]=='C4794457': L1.append(950)
    elif  l1[y]=='C4823215': L1.append(951)
    elif  l1[y]=='C4823218': L1.append(952)
    elif  l1[y]=='C4823221': L1.append(953)
    elif  l1[y]=='C4823224': L1.append(954)
    elif  l1[y]=='C4823227': L1.append(955)
    elif  l1[y]=='C4846360': L1.append(956)
    elif  l1[y]=='C4846363': L1.append(957)
    elif  l1[y]=='C4846366': L1.append(958)
    elif  l1[y]=='C4846369': L1.append(959)
    elif  l1[y]=='C4846375': L1.append(960)
    elif  l1[y]=='C4936559': L1.append(961)
    elif  l1[y]=='C4936562': L1.append(962)
    elif  l1[y]=='C4936565': L1.append(963)
    elif  l1[y]=='C4936568': L1.append(964)
    elif  l1[y]=='C4936571': L1.append(965)
    elif  l1[y]=='C4936574': L1.append(966)
    elif  l1[y]=='C4936577': L1.append(967)
    elif  l1[y]=='C4936580': L1.append(968)
    elif  l1[y]=='C4936583': L1.append(969)
    elif  l1[y]=='C4936586': L1.append(970)
    elif  l1[y]=='C4963375': L1.append(971)
    elif  l1[y]=='C4963378': L1.append(972)
    elif  l1[y]=='C4963381': L1.append(973)
    elif  l1[y]=='C4963384': L1.append(974)
    elif  l1[y]=='C4963387': L1.append(975)
    elif  l1[y]=='C4989840': L1.append(976)
    elif  l1[y]=='C4989843': L1.append(977)
    elif  l1[y]=='C4989846': L1.append(978)
    elif  l1[y]=='C4989849': L1.append(979)
    elif  l1[y]=='C4999664': L1.append(980)
    elif  l1[y]=='C4999667': L1.append(981)
    elif  l1[y]=='C4999670': L1.append(982)
    elif  l1[y]=='C4999673': L1.append(983)
    elif  l1[y]=='C4999676': L1.append(984)
    elif  l1[y]=='C4999679': L1.append(985)
    elif  l1[y]=='C5069482': L1.append(986)
    elif  l1[y]=='C5069485': L1.append(987)
    elif  l1[y]=='C5069488': L1.append(988)
    elif  l1[y]=='C5069491': L1.append(989)
    elif  l1[y]=='C5069494': L1.append(990)
    elif  l1[y]=='C5069497': L1.append(991)
    elif  l1[y]=='C5069500': L1.append(992)
    elif  l1[y]=='C5069503': L1.append(993)
    elif  l1[y]=='C5069506': L1.append(994)
    elif  l1[y]=='C5069509': L1.append(995)
    elif  l1[y]=='C5141543': L1.append(996)
    elif  l1[y]=='C5141546': L1.append(997)
    elif  l1[y]=='C5141549': L1.append(998)
    elif  l1[y]=='C5141552': L1.append(999)
    elif  l1[y]=='C5141555': L1.append(1000)
    elif  l1[y]=='C5250619': L1.append(1001)
    elif  l1[y]=='C5250622': L1.append(1002)
    elif  l1[y]=='C5250625': L1.append(1003)
    elif  l1[y]=='C5250628': L1.append(1004)
    elif  l1[y]=='C5250631': L1.append(1005)
    elif  l1[y]=='C5250634': L1.append(1006)
    elif  l1[y]=='C5250637': L1.append(1007)
    elif  l1[y]=='C5250640': L1.append(1008)
    elif  l1[y]=='C5250643': L1.append(1009)
    elif  l1[y]=='C5250646': L1.append(1010)
    elif  l1[y]=='C5250649': L1.append(1011)
    elif  l1[y]=='C5250652': L1.append(1012)
    elif  l1[y]=='C5250655': L1.append(1013)
    elif  l1[y]=='C5250658': L1.append(1014)
    elif  l1[y]=='C5250661': L1.append(1015)
    elif  l1[y]=='C5250664': L1.append(1016)
    elif  l1[y]=='C5250667': L1.append(1017)
    elif  l1[y]=='C5250670': L1.append(1018)
    elif  l1[y]=='C5250673': L1.append(1019)
    elif  l1[y]=='C5250676': L1.append(1020)
    elif  l1[y]=='C1750973': L1.append(1021)
    elif  l1[y]=='C1781505': L1.append(1022)    
    elif  l1[y]=='C5362198': L1.append(1023)  
    elif  l1[y]=='C5362201': L1.append(1024)  
    elif  l1[y]=='C5362204': L1.append(1025)  
    elif  l1[y]=='C5362207': L1.append(1026)  
    elif  l1[y]=='C5362213': L1.append(1027)  
    elif  l1[y]=='C5381815': L1.append(1028)  
    elif  l1[y]=='C5750315': L1.append(1029)
    elif  l1[y]=='C5750319': L1.append(1030)
    elif  l1[y]=='C5750326': L1.append(1031) 
    elif  l1[y]=='C6158755': L1.append(1032) 
    elif  l1[y]=='C6158760': L1.append(1033) 
    elif  l1[y]=='C6158765': L1.append(1034) 
    elif  l1[y]=='C6158770': L1.append(1035)
    elif  l1[y]=='C6158775': L1.append(1036)
    elif  l1[y]=='C6158780': L1.append(1037)
    elif  l1[y]=='C6158785': L1.append(1038)
    elif  l1[y]=='C6158790': L1.append(1039)
    elif  l1[y]=='C6158795': L1.append(1040)
    elif  l1[y]=='C6158800': L1.append(1041)
    
    else:
       L1.append(0)
       

       
for z in range(len(L1)):
    B=sheet.cell(row=z+1,column=1)
    B.value=L1[z]
    
sheet.delete_cols(2,3)    
sheet.delete_cols(5)    
wb.save(filename=""+name+"") 


from openpyxl import load_workbook
name='MonthlyBill.xlsx'
wb=load_workbook(filename=""+name+"")
sheet=wb.active
mdfi=[]
max_col=sheet.max_row

def again(t,value):
    if t=='':
        t=t+value
    else:    
        t=t+s+value
    return t

s=','
for i in range(1,max_col+1):
    cells=sheet.cell(row=i,column=6)
    mdfi.append(cells.value)
L1=[]
for i in range(len(mdfi)):
    t=''
    if mdfi[i]=='TIMES VALUE PACK, MEGA TAMIL PACK 1, TCCL BASIC FTA 170 Channel, JAYA TAMIL PACK, SUN TAMIL BASIC, ICAST TAMIL BUDGET, STAR TAMIL VALUE, DISNEY KIDS PACK, SAI TV, DISCOVERY TAMIL FAMILY PACK, SONY YAY, ZEE PRIME MOVIE PACK TAMIL SD, BBC WORLD NEWS, SONY TEN 1, TVTN NEWS BOUQUET, TCCL CLASSIC PACK 1 YEAR, SONY HAPPY INDIA ENGLISH SD, TURNER FAMILY PACK, RAJ TAMIL PACK':
        mdfi[i]='classic'
    sp=mdfi[i].split(",")
    for j in range(len(sp)):
        T=[]
        for i in range(len(sp[j])):
            T.append(sp[j][i])
   
        if T[0].isspace()==True:
            T.pop(0)
    
        if T[-1].isspace()==True:
            T.pop(-1) 
           
        c=''
        for x in T:
            c+=x
        if c=='TCCL CLASSIC SD':
            t=again(t,'classic')
        elif c=='ZEE PRIME MOVIE PACK TAMIL SD':
            t=again(t,'zee tamil')
        elif c=='TCCL BASIC FTA 170 Channel' or c=='TCCL BASIC TAMIL SD PACK':
            t=again(t,'basic')
        elif c=='TCCL TAMIL MALAYALAM SD PACK':
            t=again(t,'malayalam pack')
        elif c=='NEWS18 TAMIL NADU':
            t=again(t,'news 18')    
        elif c=='NEWS18 TAMIL NADU':
            t=again(t,'news 18')    
        elif c=='NEWS18 TAMIL NADU':
            t=again(t,'news 18')    
        else:
            t=again(t,c)
            
            
    L1.append(t)

for x in range(len(L1)):
    B=sheet.cell(row=x+1,column=6)
    B.value=L1[x]
        
wb.save(filename=""+name+"")




from openpyxl import load_workbook
name='MonthlyBill.xlsx'
wb=load_workbook(filename=""+name+"")
sheet=wb.active
mdfi1=[]
max_col=sheet.max_row


for i in range(1,max_col+1):
    cells=sheet.cell(row=i,column=5)
    mdfi1.append(cells.value)

L2=[]

for i in range(len(mdfi1)):
    s=str(mdfi1[i])
    if s[:8]=="91900000":
        L2.append("")
    else:
        if mdfi1[i]==919000000000 or mdfi1[i]==911234567890 or mdfi1[i]==910123456789 or mdfi1[i]==911234567891 or mdfi1[i]==911234657890:
            L2.append("")
        else:
            L2.append(mdfi1[i])
            

for x in range(len(L2)):
    B=sheet.cell(row=x+1,column=5)
    B.value=L2[x]

wb.save(filename=""+name+"")

print("  "*8,"MISSION COMPLETED")